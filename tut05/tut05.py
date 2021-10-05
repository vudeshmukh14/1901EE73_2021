import csv
import os
import openpyxl

def generate_marksheet(names_roll, subjects_master, grade_numeric_equivalent):

    with open("grades.csv") as f:  #opening grades.csv file in read mode
        reader=csv.reader(f)

        last_fileName="None"   #initialising variables which will be required in calculations
        sem_credit, spi, total_credit, curr_sum, last_sem, last_col=0, 0, 0, 0, 0, 2

        for new_row in reader: #reading each row one by one in reader object
            if new_row[0]=="Roll":  #skipping the first row as it contains headers which we dont require
                continue

            fileName=new_row[0]+".xlsx"    #generating file name using roll number
            if os.path.exists(os.path.join("output", fileName))==False:   #if the file doesn't already exist then we create it

                if last_fileName!="None":  #this if block calculates CPI, SPI etc for the last marksheet before current one
                    wb=openpyxl.load_workbook(os.path.join("output", last_fileName))
                    sheet=wb["Overall"]
                    sheet["B1"]=last_fileName[0:8]
                    sheet["B2"]=names_roll[last_fileName[0:8]]
                    sheet["B3"]=last_fileName[4:6]
                    spi=spi/sem_credit
                    temp_sheet=wb["Overall"]
                    temp_sheet.cell(row=5, column=last_col).value=sem_credit
                    temp_sheet.cell(row=6, column=last_col).value=round(spi, 2)
                    total_credit=total_credit+sem_credit
                    curr_sum=curr_sum+spi*sem_credit
                    temp_sheet.cell(row=7, column=last_col).value=total_credit
                    temp_sheet.cell(row=8, column=last_col).value=round(curr_sum/total_credit, 2)
                    wb.save(os.path.join("output", last_fileName))
                    spi, sem_credit, last_sem, last_col, curr_sum, total_credit=0, 0, int(new_row[1]), 2, 0, 0

                last_fileName=fileName
                wb=openpyxl.Workbook()  #opening a new workbook and creating required sheets
                title_change=wb["Sheet"]
                title_change.title="Overall"
                wb.create_sheet(index=1, title="Sem1")
                wb.create_sheet(index=2, title="Sem2")
                wb.create_sheet(index=3, title="Sem3")
                wb.create_sheet(index=4, title="Sem4")
                wb.create_sheet(index=5, title="Sem5")
                wb.create_sheet(index=6, title="Sem6")
                wb.create_sheet(index=7, title="Sem7")
                wb.create_sheet(index=8, title="Sem8")

                sheet=wb["Overall"]   #filling common values in Overall sheet
                sheet.cell(row=1, column=1).value="Roll No."
                sheet.cell(row=2, column=1).value="Name of Student"
                sheet.cell(row=3, column=1).value="Discipline"
                sheet.cell(row=4, column=1).value="Semester No."
                sheet.cell(row=5, column=1).value="Semester wise Credit Taken"
                sheet.cell(row=6, column=1).value="SPI"
                sheet.cell(row=7, column=1).value="Total Credits Taken"
                sheet.cell(row=8, column=1).value="CPI"
                sheet.cell(row=4, column=2).value=1
                sheet.cell(row=4, column=3).value=2
                sheet.cell(row=4, column=4).value=3
                sheet.cell(row=4, column=5).value=4
                sheet.cell(row=4, column=6).value=5
                sheet.cell(row=4, column=7).value=6
                sheet.cell(row=4, column=8).value=7
                sheet.cell(row=4, column=9).value=8
                first_row=("Sl No.", "Subject No.", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade")
                for x in range(1, 9):
                    sheet=wb["Sem"+str(x)]
                    sheet.append(first_row)
                
                sheet=wb["Sem"+str(new_row[1])] #appending the current row and calculating required things from this row
                sheet.append((1, new_row[2], subjects_master[new_row[2]][0], subjects_master[new_row[2]][1], int(subjects_master[new_row[2]][2]), new_row[5], new_row[4]))
                last_sem=1
                sem_credit=sem_credit+int(new_row[3])
                spi=spi+int(new_row[3])*grade_numeric_equivalent[new_row[4].strip()]
                wb.save(os.path.join("output", fileName))  #finally saving the workbook

            else :   #when the excel file with the name fileName already exists
                wb=openpyxl.load_workbook(os.path.join("output", fileName))

                if ("Sem"+str(new_row[1])) not in wb.sheetnames:  #for handling corner case if there are more than 8 semesters
                    wb.create_sheet("Sem"+str(new_row[1]))        #found such case for 1401ME11, 10th semester was found
                    sheet=wb["Sem"+str(new_row[1])]
                    first_row=("Sl No.", "Subject No.", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade")
                    sheet.append(first_row)
                    sheet=wb["Overall"]
                    sheet.cell(row=4, column=sheet.max_column+1).value=int(new_row[1])

                sheet=wb["Sem"+str(new_row[1])]  #appending the current row in appropriate sheet
                sheet.append((sheet.max_row, new_row[2], subjects_master[new_row[2]][0], subjects_master[new_row[2]][1], int(subjects_master[new_row[2]][2]), new_row[5], new_row[4]))
                
                if last_sem!=int(new_row[1]):    #if this is the new sem beginning, then calculating CPI, SPI etc for last sem
                    spi=spi/sem_credit
                    temp_sheet=wb["Overall"]
                    temp_sheet.cell(row=5, column=last_col).value=sem_credit
                    temp_sheet.cell(row=6, column=last_col).value=round(spi, 2)
                    total_credit=total_credit+sem_credit
                    curr_sum=curr_sum+spi*sem_credit
                    temp_sheet.cell(row=7, column=last_col).value=total_credit
                    temp_sheet.cell(row=8, column=last_col).value=round(curr_sum/total_credit, 2)
                    spi, sem_credit, last_sem, last_col=0, 0, int(new_row[1]), last_col+1

                sem_credit=sem_credit+int(new_row[3])
                spi=spi+int(new_row[3])*grade_numeric_equivalent[new_row[4].strip()]
                wb.save(os.path.join("output", fileName))

        #the block of code below is for the last excel file for which calulations
        #will not be done in the above loop
        wb=openpyxl.load_workbook(os.path.join("output", last_fileName))
        sheet=wb["Overall"]
        sheet["B1"]=last_fileName[0:8]
        sheet["B2"]=names_roll[last_fileName[0:8]]
        sheet["B3"]=last_fileName[4:6]
        spi=spi/sem_credit
        temp_sheet=wb["Overall"]
        temp_sheet.cell(row=5, column=last_col).value=sem_credit
        temp_sheet.cell(row=6, column=last_col).value=round(spi, 2)
        total_credit=total_credit+sem_credit
        curr_sum=curr_sum+spi*sem_credit
        temp_sheet.cell(row=7, column=last_col).value=total_credit
        temp_sheet.cell(row=8, column=last_col).value=round(curr_sum/total_credit, 2)
        wb.save(os.path.join("output", last_fileName))

    return



names_roll={}    #mapping a dictionary for names-roll.csv file with roll no as key and name as value 
with open("names-roll.csv", "r") as f:
    reader=csv.reader(f)
    for row in reader:
        names_roll[row[0]]=row[1]

subjects_master={} #mapping a dictionary for subject_master.csv file with subno as key and list of subname, ltp and crd as value
with open("subjects_master.csv", "r") as f:
    reader=csv.reader(f)
    for row in reader:
        subjects_master[row[0]]=[row[1], row[2], row[3]]

grade_numeric_equivalent={  #mapping a dictionary from grades to their numerical equivalent
    "AA":10,
    "AA*":10,
    "AB":9,
    "AB*":9,
    "BB":8,
    "BB*":8,
    "BC":7,
    "BC*":7,
    "CC":6,
    "CC*":6,
    "CD":5,
    "CD*":5,
    "DD":4,
    "DD*":4,
    "F":0,
    "F*":0,
    "I":0,
    "I*":0
}

if os.path.exists("output")==False:  #checking if output folder exists in the same directory
    os.mkdir("output")               #if not, then creating output folder
  
generate_marksheet(names_roll, subjects_master, grade_numeric_equivalent)  #calling function