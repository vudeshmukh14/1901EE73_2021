from flask import Flask, render_template, request
import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
import os
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from string import Template

# making required imports

# initializing Flask object
app = Flask(__name__, template_folder="templates")
# setting path for uploaded files to be saved
app.config['UPLOAD_FOLDER']='D:\\1901EE73_2021\\proj1'

@app.route("/")
def marksheet_generator():
    return render_template("index.html")


@app.route('/upload', methods = ['POST', 'GET'])  # for uploading all the files
def upload():
    if request.method == 'POST':
        f = request.files['File']
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], os.path.join("preprocessing", f.filename)))
        return render_template("upload.html", message="File saved successfully, you can go back :)")


@app.route('/', methods=['POST'])  # taking input for positive and negative markings
def marking_scheme():
    text = request.form['text']
    processed_text = text.upper()
    if int(processed_text)>0:
        f=open("preprocessing\\pos_marks.txt", "w")
        f.write(processed_text)
        f.close()
    else :
        f=open("preprocessing\\neg_marks.txt", "w")
        f.write(processed_text)
        f.close()
    return "Input taken, you can go back :)"


def generate_marksheets():  # for generating marksheets for all roll numbers

    # if marksheets folder doesn't exist then creating it
    if os.path.exists("marksheets")==False:
        os.mkdir("marksheets")

    dict_responses={} # {roll: [marked options, ]}

    # checking if the required files are uploaded or not
    if os.path.exists("preprocessing\\responses.csv")==False:
        return render_template("forward.html", message="Please upload responses.csv")
    if os.path.exists("preprocessing\\master_roll.csv")==False:
        return render_template("forward.html", message="Please upload master_roll.csv")
    if os.path.exists("preprocessing\\pos_marks.txt")==False:
        return render_template("forward.html", message="Please enter marks for correct answer")
    if os.path.exists("preprocessing\\neg_marks.txt")==False:
        return render_template("forward.html", message="Please enter marks for correct answer")

    # creating dictionary for easy and fast access
    with open("preprocessing\\responses.csv", "r") as f:
        reader=csv.reader(f)
        for row in reader:
            if row[0]=="Timestamp":
                continue
            dict_responses[row[6]]=[row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14], row[15], row[16], row[17], row[18], row[19], row[20], row[21], row[22], row[23], row[24], row[25], row[26], row[27], row[28], row[29], row[30], row[31], row[32], row[33], row[34]]

    with open("preprocessing\\pos_marks.txt", "r") as f:
        line=f.readline()
        pos_marks=int(line.strip())

    with open("preprocessing\\neg_marks.txt", "r") as f:
        line=f.readline()
        neg_marks=int(line.strip())

    if "ANSWER" in dict_responses.keys()==False:
        return render_template("forward.html", message="Answer key not found :(")
    
    # iterating for all roll numbers
    with open("preprocessing\\master_roll.csv", "r") as f:
        reader=csv.reader(f)

        for row in reader:

            if row[0]=="roll":
                continue

            # variables for calculations
            correct=0
            wrong=0
            unattempted=0

            # counting number of right, wrong and unattempted answers
            for x in range(28):
                if row[0] in dict_responses.keys()==False:
                    dict_responses[row[0]]=["","","","","","","","","","","","","","","","","","","","","","","","","","","",""]
                    unattempted+=1
                elif dict_responses[row[0]][x]=="":
                    unattempted+=1
                elif dict_responses["ANSWER"][x]==dict_responses[row[0]][x]:
                    correct+=1
                else :
                    wrong+=1

            # creating excel file
            wb=Workbook()
            sheet=wb.active

            # increasing default column widths and row heights for better viewing
            sheet.column_dimensions['A'].width = 23.7
            sheet.column_dimensions['B'].width = 23.7
            sheet.column_dimensions['C'].width = 23.7
            sheet.column_dimensions['D'].width = 23.7
            sheet.column_dimensions['E'].width = 23.7
            sheet.row_dimensions[1].height = 25
            sheet.row_dimensions[2].height = 25
            sheet.row_dimensions[3].height = 25
            sheet.row_dimensions[4].height = 25

            # marksheet summary
            sheet.cell(row=9, column=1).value=""
            sheet.cell(row=9, column=2).value="Right"
            sheet.cell(row=9, column=3).value="Wrong"
            sheet.cell(row=9, column=4).value="Not Attempt"
            sheet.cell(row=9, column=5).value="Max"
            sheet.cell(row=10, column=1).value="No."
            sheet.cell(row=11, column=1).value="Marking"
            sheet.cell(row=12, column=1).value="Total"
            sheet.cell(row=10, column=2).value=correct
            sheet.cell(row=10, column=3).value=wrong
            sheet.cell(row=10, column=4).value=unattempted
            sheet.cell(row=10, column=5).value=28
            sheet.cell(row=11, column=2).value=pos_marks
            sheet.cell(row=11, column=3).value=neg_marks
            sheet.cell(row=11, column=4).value=0
            sheet.cell(row=11, column=5).value=""
            sheet.cell(row=12, column=2).value=correct*pos_marks
            sheet.cell(row=12, column=3).value=wrong*neg_marks
            sheet.cell(row=12, column=4).value=""
            sheet.cell(row=12, column=5).value=(str((correct*pos_marks+wrong*neg_marks))+"/140")

            # borders and styling
            rows=sheet["A9:E12"]
            for r in rows:
                for cell in r:
                    cell.border=Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
            ftred = Font(color="FF0000")
            ftgreen = Font(color="008000")
            ftblue = Font(color="0000FF")
            sheet['C10'].font=ftred
            sheet['C11'].font=ftred
            sheet['C12'].font=ftred
            sheet['E12'].font=ftblue
            sheet['B10'].font=ftgreen
            sheet['B11'].font=ftgreen
            sheet['B12'].font=ftgreen
            sheet['B9'].font=Font(bold=True)
            sheet['C9'].font=Font(bold=True)
            sheet['D9'].font=Font(bold=True)
            sheet['E9'].font=Font(bold=True)
            sheet['A10'].font=Font(bold=True)
            sheet['A11'].font=Font(bold=True)
            sheet['A12'].font=Font(bold=True)

            # header section (logo and student information)
            img = openpyxl.drawing.image.Image('logo.png')
            img.anchor = 'A1'
            sheet.add_image(img)
            sheet.merge_cells('A5:E5')
            sheet.cell(row=5, column=1).value="Mark Sheet"
            sheet['A5'].font=Font(bold=True)
            sheet['A5'].font=Font(size=24)
            sheet['A5'].alignment = Alignment(horizontal='center')
            sheet['A6'].value="Name:"
            sheet.merge_cells("B6:C6")
            sheet['B6'].value=row[1]
            sheet['B6'].font=Font(bold=True)
            sheet['D6'].value="Exam:"
            sheet['E6'].value="quiz"
            sheet['E6'].font=Font(bold=True)
            sheet['A7'].value="Roll Number:"
            sheet['B7'].value=row[0]
            sheet['B7'].font=Font(bold=True)
            sheet['A15'].value="Student Ans"
            sheet['A15'].font=Font(bold=True)
            sheet['B15'].value="Correct Ans"
            sheet['B15'].font=Font(bold=True)
            sheet['D15'].value="Student Ans"
            sheet['D15'].font=Font(bold=True)
            sheet['E15'].value="Correct Ans"
            sheet['E15'].font=Font(bold=True)

            # all the marked options of student and correct answers
            for x in range(16, 44):
                if x>40:
                    sheet.cell(row=x-25, column=5).value=dict_responses["ANSWER"][x-16]
                    sheet.cell(row=x-25, column=5).font=ftblue
                    continue
                sheet.cell(row=x, column=2).value=dict_responses["ANSWER"][x-16]
                sheet.cell(row=x, column=2).font=ftblue
            for x in range(16, 44):
                if x>40:
                    sheet.cell(row=x-25, column=4).value=dict_responses[row[0]][x-16]
                    if dict_responses[row[0]][x-16]==dict_responses["ANSWER"][x-16]:
                        sheet.cell(row=x-25, column=4).font=ftgreen
                    else :
                        sheet.cell(row=x-25, column=4).font=ftred
                    continue
                sheet.cell(row=x, column=1).value=dict_responses[row[0]][x-16]
                if dict_responses[row[0]][x-16]==dict_responses["ANSWER"][x-16]:
                    sheet.cell(row=x, column=1).font=ftgreen
                else :
                    sheet.cell(row=x, column=1).font=ftred

            # border and styling
            rows=sheet["A15:B40"]
            for r in rows:
                for cell in r:
                    cell.border=Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
            rows=sheet["D15:E18"]
            for r in rows:
                for cell in r:
                    cell.border=Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
                
            sheet.title="quiz"
            wb.save(os.path.join("marksheets", str(row[0])+".xlsx"))

    return render_template("forward.html", message="Generated Roll number wise Marksheets, you can go back now :)")


def concise_marksheets():
    # creating the excel file
    wb2=Workbook()
    sheet=wb2.active
    sheet.title="concise_marksheet"

    # list for answer key
    correct_ans=[]

    # checking if the required files exist or not
    if os.path.exists("preprocessing\\responses.csv")==False:
        return render_template("forward.html", message="Please upload responses.csv")
    if os.path.exists("preprocessing\\master_roll.csv")==False:
        return render_template("forward.html", message="Please upload master_roll.csv")
    if os.path.exists("preprocessing\\pos_marks.txt")==False:
        return render_template("forward.html", message="Please enter marks for correct answer")
    if os.path.exists("preprocessing\\neg_marks.txt")==False:
        return render_template("forward.html", message="Please enter marks for correct answer")

    with open("preprocessing\\pos_marks.txt", "r") as f:
        line=f.readline()
        pos_marks=int(line.strip())

    with open("preprocessing\\neg_marks.txt", "r") as f:
        line=f.readline()
        neg_marks=int(line.strip())

    # finding the answer key and storing it in the list
    with open("preprocessing\\responses.csv", "r") as file:
        reader=csv.reader(file)
        for row in reader:
            if row[6]=="ANSWER":
                for x in range(0, 28):
                    correct_ans.append(row[7+x])
                break
    
    if len(correct_ans)==0:
        return render_template("forward.html", message="Answer key not found :(")

    # iterating over responses.csv
    with open("preprocessing\\responses.csv", "r") as file:
        reader=csv.reader(file)
        for row in reader:
            if row[0]=="Timestamp":
                row[2]="Google_score"
                row.insert(6, "Score_After_Negative")
                row.append("statusAns")
                sheet.append(row)
                continue

            # required variables for calculation
            right=0
            wrong=0
            unattempted=0

            # counting
            for x in range(7, 35):
                if row[x]=="":
                    unattempted+=1
                elif row[x]==correct_ans[x-7]:
                    right+=1
                else :
                    wrong+=1

            row[2]=str(right*pos_marks)+" / 140"
            row.insert(6, str(right*pos_marks+wrong*neg_marks)+" / 140")
            row.append("["+str(right)+", "+str(wrong)+", "+str(unattempted)+"]")
            sheet.append(row)

    wb2.save(os.path.join("marksheets", "concise_marksheet.xlsx"))

    return render_template("forward.html", message="Generated Concise Marksheets, you can go back now :)")


def read_template(filename):  # function te read message body and converting to a template

    with open(filename, 'r', encoding='utf-8') as template_file:
        template_file_content = template_file.read()
    return Template(template_file_content)


def send_emails():  # function for sending emails

    dict_responses={}

    if os.path.exists("preprocessing\\responses.csv")==False:
        return render_template("forward.html", message="Please generate marksheets first")
    if os.path.exists("preprocessing\\master_roll.csv")==False:
        return render_template("forward.html", message="Please generate marksheets first")

    with open("preprocessing\\responses.csv", "r") as file:
        reader=csv.reader(file)
        for row in reader:
            if row[0]=="Timestamp":
                continue
            dict_responses[row[6]]=[row[1], row[4]]

    # iterating for all roll numbers
    with open("preprocessing\\master_roll.csv", "r") as file:
        reader=csv.reader(file)

        for contact in reader:
            if contact[0]=="roll":
                continue

            fromaddr = "skn.frost@gmail.com"
            toaddr1 = dict_responses[contact[0]][0].strip()
            toaddr2 = dict_responses[contact[0]][1].strip()

            # sending emails to gmail ids
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr1
            msg['Subject'] = "CS384 2021 - quiz - with Negative"
            message_template = read_template("preprocessing\\message.txt")
            message = message_template.substitute(PERSON_NAME=contact[1].title())
            msg.attach(MIMEText(message, 'plain'))
            filename = str(contact[0])+".xlsx"
            if os.path.exists("marksheets")==False:
                return render_template("forward.html", message="Please generate marksheets first")
            if os.path.exists(os.path.join("marksheets", filename))==False:
                return render_template("forward.html", message="Please generate marksheets first")
            attachment = open(os.path.join("marksheets", filename), "rb")
            p = MIMEBase('application', 'octet-stream')
            p.set_payload((attachment).read())
            encoders.encode_base64(p)
            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
            msg.attach(p)
            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login(fromaddr, "12121212sumit")
            text = msg.as_string()
            s.sendmail(fromaddr, toaddr1, text)
            s.quit()

            # sending emails to webmail ids
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr2
            msg['Subject'] = "CS384 2021 - quiz - with Negative"
            message_template = read_template("preprocessing\\message.txt")
            message = message_template.substitute(PERSON_NAME=contact[1].title())
            msg.attach(MIMEText(message, 'plain'))
            filename = str(contact[0])+".xlsx"
            attachment = open(os.path.join("marksheets", filename), "rb")
            p = MIMEBase('application', 'octet-stream')
            p.set_payload((attachment).read())
            encoders.encode_base64(p)
            p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
            msg.attach(p)
            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login(fromaddr, "12121212sumit")
            text = msg.as_string()
            s.sendmail(fromaddr, toaddr2, text)
            s.quit()

    return render_template("forward.html", message="E-mails sent, you can go back :)")


@app.route("/forward", methods=['POST'])  # action taken on button clicks
def button_clicks():
    if request.method == 'POST':
        if request.form['submit_button'] == 'Roll number wise':
            return generate_marksheets()
        elif request.form['submit_button'] == 'Concise':
            return concise_marksheets()
        elif request.form['submit_button'] == 'Emails':
           return send_emails()

if __name__=="__main__":
    app.run(debug=True)

