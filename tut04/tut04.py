import os
import csv
from openpyxl import Workbook, load_workbook

def output_by_subject(file):
    if os.path.isdir("output_by_subject")==False:
        os.mkdir("output_by_subject")

    reader=csv.reader(file, delimiter=',', skipinitialspace=True)
    first=True
    header=[]
    for row in reader:
        if first:
            header=[row[0]+row[1]+row[3]+row[8]]
            first=False
        else :
            if os.path.exists("output_by_subject/%s.xlsx"%row[3])==False:
                wb=Workbook()
                sheet=wb.active
                sheet.append(header)
                wb.save("output_by_subject/%s.xlsx"%row[3])
            
            wb=load_workbook("output_by_subject/%s.xlsx"%row[3])
            sheet=wb.active
            sheet.append([row[0]+row[1]+row[3]+row[8]])
            wb.save("output_by_subject/%s.xlsx"%row[3])
    return

def output_individual_roll(file):
    if os.path.isdir("output_individual_roll")==False:
        os.mkdir("output_individual_roll")

    reader=csv.reader(file, delimiter=',', skipinitialspace=True)
    first=True
    header=[]
    for row in reader:
        if first:
            header=[row[0]+row[1]+row[3]+row[8]]
            first=False
        else :
            if os.path.exists("output_individual_roll/%s.xlsx"%row[0])==False:
                wb=Workbook()
                sheet=wb.active
                sheet.append(header)
                wb.save("output_individual_roll/%s.xlsx"%row[0])
            
            wb=load_workbook("output_individual_roll/%s.xlsx"%row[0])
            sheet=wb.active
            sheet.append([row[0]+row[1]+row[3]+row[8]])
            wb.save("output_individual_roll/%s.xlsx"%row[0])

    return

with open("regtable_old.csv") as file:
    output_individual_roll(file)

with open("regtable_old.csv") as file:
    output_by_subject(file)
